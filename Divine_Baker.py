# This example requires the 'message_content' intent.

import discord
import logging
from discord.ext import commands
from openpyxl import load_workbook

workbook = load_workbook('Bread List.xlsx')
sheet = workbook.active
breadTypes = []

handler = logging.FileHandler(filename='discord.log', encoding='utf-8', mode='w')


intents = discord.Intents.default()
intents.message_content = True
intents.members = True

bot = commands.Bot(command_prefix='!', intents=intents)

client = discord.Client(intents=intents)

def get_rand_breadname():
    pass

def contains_elem(list, elem):
    return (elem in list)

@client.event
async def on_ready():
    # Iterate through columns from 'A' to 'C' (min_col=1, max_col=3)
    # You can also specify min_row and max_row for a specific range within the columns
    for col_cells in sheet.iter_cols(min_col=0, max_col=1, values_only=True):
        for cell_value in col_cells:
            # print(cell_value)
            breadTypes.append(cell_value)

    print(f'We have logged in as {client.user}')
    for guild in client.guilds:
        inUseNames = []
        availableNames = []
        print(f"Guild: {guild.name} (ID: {guild.id})")
        print("Members:")
        for member in guild.members:
            # print(member.name, client.user.name)
            # add current nicknames
            if member.nick != None:
                inUseNames.append(member.nick)
            
            if member.nick == None:
                await setnick(commands.context, member, "okayLilBro")
            # print(f"  - {member.name} (ID: {member.id})")

@client.event
async def on_message(message):
    if message.author == client.user:
        return

    if message.content.startswith('$hello'):
        await message.channel.send('Hello!')

@bot.command()
@commands.has_permissions(manage_nicknames=True) # Ensure the command caller has permissions
async def setnick(ctx, member: discord.Member, new_nickname: str):
    # """Changes the nickname of a specified member."""
    try:
        await member.edit(nick=new_nickname)
        # await ctx.send(f'Successfully changed {member.display_name}\'s nickname to {new_nickname}.')
        return
    except discord.Forbidden:
        # await ctx.send("I don't have permission to change that user's nickname.")
        return
    except discord.HTTPException as e:
        # await ctx.send(f"An error occurred while changing the nickname: {e}")
        return

client.run('xxx', log_handler=handler)